import pandas as pd
import shutil
from openpyxl import load_workbook
from datetime import datetime


class SpecificationGenerator:

    def __init__(self, master_file_path, fort_qr_path, template_path):
        self.master_file_path = master_file_path
        self.fort_qr_path = fort_qr_path
        self.template_path = template_path
        self.output_df = pd.DataFrame(columns=[
            'productNameRus',
            'productNameEng',
            'identificationCode',
            'identificationCodeOuter',
            'identificationCodeCase',
            'identificationCodePallet',
            'invoiceNo',
            'invoiceDate',
            'TotalAmount'
        ])
        self.new_filename = None

    def load_data(self):
        # Читаем данные из XLSX файлов
        # Для Мастер файла: заголовки в 1-й строке (индекс 0), данные со 2-й (индекс 1)
        master_file = pd.read_excel(self.master_file_path, sheet_name=0, header=0)
        # Ограничиваем master_file до строк, где SIZE5 не пустой
        master_file = master_file[master_file['SIZE5'].notna()]

        # Для FORT_QR: заголовки в 7-й строке (индекс 6), данные с 11-й (индекс 10)
        fort_qr = pd.read_excel(self.fort_qr_path, sheet_name=0, header=6, skiprows=range(7, 10))

        # Преобразуем GTIN и GTIN Outer в строки, если они не пустые
        master_file['GTIN'] = master_file['GTIN'].apply(lambda x: str(int(x)) if pd.notna(x) else x)
        master_file['GTIN Outer'] = master_file['GTIN Outer'].apply(lambda x: str(int(x)) if pd.notna(x) else x)
        fort_qr['GTIN'] = fort_qr['GTIN'].apply(lambda x: str(int(x)) if pd.notna(x) else x)

        return master_file, fort_qr

    def process_data(self, master_file, fort_qr):
        # 1. Находим все GTIN Outer из мастер файла в FORT_QR
        # outer_matches = fort_qr[fort_qr['GTIN'].isin(master_file['GTIN Outer'])]

        # --------------------------------------------------------------
        # 1. Делаем вспомогательный DataFrame только с нужными колонками из мастер-файла
        #    (SKU → GTIN Outer). Если один SKU встречается несколько раз — оставляем все строки.
        # --------------------------------------------------------------
        master_lookup = master_file[['SKU', 'GTIN Outer']].dropna().copy()

        # Приводим типы к строке (на всякий случай, чтобы не было проблем с int/float)
        master_lookup['SKU'] = master_lookup['SKU'].astype(str)
        master_lookup['GTIN Outer'] = master_lookup['GTIN Outer'].astype(str)

        fort_qr['buyerSKU'] = fort_qr['buyerSKU'].astype(str)
        fort_qr['GTIN'] = fort_qr['GTIN'].astype(str)

        # --------------------------------------------------------------
        # 2. Присоединяем к fort_qr информацию о том, есть ли такое сочетание SKU + GTIN Outer в мастер-файле
        #    how='left' — оставляем все строки из fort_qr
        #    left_on  — две колонки слева
        #    right_on — две колонки справа
        # --------------------------------------------------------------
        merged = fort_qr.merge(
            master_lookup,
            how='left',
            left_on=['buyerSKU', 'GTIN'],
            right_on=['SKU', 'GTIN Outer']
        )

        # --------------------------------------------------------------
        # 3. Оставляем только те строки fort_qr, у которых нашлось совпадение
        #    (т.е. колонка 'SKU' не NaN после merge)
        # --------------------------------------------------------------
        outer_matches = merged[merged['SKU'].notna()].copy()

        # Если нужно — можно сразу удалить вспомогательные колонки, чтобы дальше работать как раньше
        outer_matches = outer_matches.drop(columns=['SKU', 'GTIN Outer_y'], errors='ignore')
        # (GTIN Outer может быть и под именем 'GTIN Outer_x' — оставляем оригинальный 'GTIN')

        # print(f"Найдено совпадений по GTIN Outer + buyerSKU: {len(outer_matches)}")
        # --------------------------------------------------------------

        # 2. Считаем количество совпадений
        outer_count = len(outer_matches)
        print(f"Найдено совпадений GTIN Outer: {outer_count}")

        # 3. Итерация по мастер файлу
        row_index = 0

        for _, master_row in master_file.iterrows():
            current_gtin = master_row['GTIN Outer']
            print(f"\nОбработка GTIN Outer: {current_gtin}")

            # Находим все строки с соответствующим GTIN
            # pack_rows = fort_qr[fort_qr['GTIN'] == master_row['GTIN']]
            # ------------------------------------------------------------------
            # Ищем пачки (GTIN пачки), которые:
            # 1. Имеют GTIN == master_row['GTIN'] (это код пачки)
            # 2. И при этом их buyerSKU соответствует SKU из текущей строки master_file
            # ------------------------------------------------------------------
            current_sku = str(master_row['SKU']).strip()
            current_gtin_pack = str(master_row['GTIN']).strip() if pd.notna(master_row['GTIN']) else None

            if not current_gtin_pack or current_gtin_pack == 'nan':
                print(f"Предупреждение: у SKU {current_sku} пустой GTIN пачки — пропускаем")
                continue  # или можно задать пустой pack_rows

            # Фильтруем только пачки с правильным GTIN И правильным buyerSKU
            pack_rows = fort_qr[
                (fort_qr['GTIN'] == current_gtin_pack) &
                (fort_qr['buyerSKU'].astype(str).str.strip() == current_sku)
                ]

            # Дополнительно: если вдруг ничего не нашли — выводим предупреждение
            if pack_rows.empty:
                print(f"Внимание: для SKU {current_sku} | GTIN пачки {current_gtin_pack} "
                      f"не найдено ни одной пачки с buyerSKU = {current_sku}")

            # Ограничиваем количество строк согласно SIZE5
            size5 = int(master_row['SIZE5']) if pd.notna(master_row['SIZE5']) else float('inf')

            # Рассчитываем количество полных порций
            total_rows = len(pack_rows)
            full_chunks = total_rows // size5
            remainder = total_rows % size5

            print(
                f"Всего пачек: {total_rows}, размер коробки (SIZE5): {size5}, "
                f"полных коробок: {full_chunks}, остаток пачек: {remainder}"
            )
            # Проверяем наличие остатка
            if remainder > 0:
                error_msg = (f"Ошибка распределения для GTIN Outer {current_gtin}: "
                             f"{total_rows} пачек не могут быть равномерно распределены "
                             f"в коробки по {size5} пачек. Остаток: {remainder} пачек.")
                raise ValueError(error_msg)

            # Обрабатываем полные порции
            for chunk_num in range(full_chunks):
                start_idx = chunk_num * size5
                end_idx = (chunk_num + 1) * size5
                chunk = pack_rows.iloc[start_idx:end_idx]

                # print(f"Обработка порции {chunk_num + 1}/{full_chunks}: строки {start_idx + 1}-{end_idx}")
                # Получаем identificationCodeOuter для текущего GTIN Outer
                outer_rows = fort_qr[fort_qr['GTIN'] == master_row['GTIN Outer']]
                identification_outer = outer_rows['identificationCode'].iloc[chunk_num] if not outer_rows.empty else None

                for _, pack_row in chunk.iterrows():
                    self.output_df.loc[row_index] = {
                        'productNameRus': pack_row['productNameRus'],
                        'productNameEng': pack_row['productNameEng'],
                        'identificationCode': pack_row['identificationCode'],
                        'identificationCodeOuter': identification_outer,
                        'identificationCodeCase': None,
                        'identificationCodePallet': None,
                        'invoiceNo': None,
                        'invoiceDate': None,
                        'TotalAmount': None
                    }
                    row_index += 1

            print(f"Обработано строк для GTIN Outer {current_gtin}: {row_index}")

        return row_index

    def save_to_excel(self, row_index):
        # Генерируем уникальное имя для нового файла с временной меткой
        self.new_filename = self.new_specification_file()

        # Загружаем новый файл для записи данных
        wb = load_workbook(self.new_filename)
        ws = wb['Invoice specification']

        # Записываем новые данные, начиная с 11-й строки (индекс 10 в Python)
        for index, row in self.output_df.iterrows():
            excel_row = 11 + index
            ws.cell(row=excel_row, column=1, value=row['productNameRus'])
            ws.cell(row=excel_row, column=2, value=row['productNameEng'])
            ws.cell(row=excel_row, column=3, value=row['identificationCode'])
            ws.cell(row=excel_row, column=4, value=row['identificationCodeOuter'])
            ws.cell(row=excel_row, column=5, value=row['identificationCodeCase'])
            ws.cell(row=excel_row, column=6, value=row['identificationCodePallet'])
            ws.cell(row=excel_row, column=7, value=row['invoiceNo'])
            ws.cell(row=excel_row, column=8, value=row['invoiceDate'])
            ws.cell(row=excel_row, column=9, value=row['TotalAmount'])

        # Сохраняем изменения в новый файл
        wb.save(self.new_filename)
        wb.close()

        print(f"Обработано строк: {row_index}")
        print(f"Результат сохранен в '{self.new_filename}', начиная с 11-й строки")

    def new_specification_file(self):
        # Генерируем уникальное имя для нового файла с временной меткой
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        new_filename = f"Invoice_Specification_{timestamp}.xlsx"

        # Копируем шаблон в новый файл
        shutil.copy(self.template_path, new_filename)

        return new_filename

    def run(self):
        # Основной метод для выполнения всех шагов
        try:
            master_file, fort_qr = self.load_data()
            row_index = self.process_data(master_file, fort_qr)
            self.save_to_excel(row_index)
        except ValueError as e:
            print(f"\nОшибка: {e}")
            print("Процесс остановлен из-за несоответствия количества пачек и коробок.")
            raise


# Вызов класса
if __name__ == "__main__":
    generator = SpecificationGenerator(
        master_file_path="Мастер файл номенклатуры.xlsx",
        fort_qr_path="Коды маркировки № 695 от 15.07.2025.xlsx",
        template_path="Invoice Specification template.xlsx"
    )
    generator.run()
