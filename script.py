import pandas as pd
import shutil
from openpyxl import load_workbook
from datetime import datetime


class SpecificationGenerator:
    """
    Генератор спецификаций для обработки данных о товарах и QR-кодах маркировки.

    Класс предназначен для создания спецификации на основе данных из мастер-файла
    номенклатуры и файла QR-кодов FORT, с последующим сохранением результата
    в файл Excel по заданному шаблону.

    Основные функции:
    - Сопоставление кодов GTIN между мастер-файлом и файлом QR-кодов
    - Группировка товаров по иерархии: пачка → коробка → мастер-кейс → паллет
    - Валидация корректности распределения товаров по упаковкам
    - Формирование итоговой спецификации с кодами идентификации
    """

    # Константы для структуры выходного файла
    OUTPUT_COLUMNS = [
        'productNameRus', 'productNameEng', 'identificationCode',
        'identificationCodeOuter', 'identificationCodeCase', 'identificationCodePallet',
        'invoiceNo', 'invoiceDate', 'TotalAmount'
    ]

    # Параметры для чтения файлов
    MASTER_FILE_HEADER_ROW = 0
    FORT_QR_HEADER_ROW = 6
    FORT_QR_SKIP_ROWS = range(7, 10)
    OUTPUT_START_ROW = 11

    def __init__(self, master_file_path: str, fort_qr_path: str, template_path: str):
        """
        Инициализация генератора спецификаций.

        Args:
            master_file_path: Путь к мастер-файлу номенклатуры (.xlsx)
            fort_qr_path: Путь к файлу с QR-кодами FORT (.xlsx)
            template_path: Путь к шаблону спецификации (.xlsx)
        """
        self.master_file_path = master_file_path
        self.fort_qr_path = fort_qr_path
        self.template_path = template_path
        self.output_df = pd.DataFrame(columns=self.OUTPUT_COLUMNS)
        self.new_filename = None

    def load_data(self) -> tuple[pd.DataFrame, pd.DataFrame]:
        """
        Загрузка и предварительная обработка данных из Excel файлов.

        Returns:
            Кортеж из DataFrame'ов (master_file, fort_qr)
        """
        # Загрузка мастер-файла
        master_file = self._load_master_file()

        # Загрузка файла FORT_QR
        fort_qr = self._load_fort_qr_file()

        # Преобразование GTIN кодов в строковый формат
        self._convert_gtin_to_string(master_file, fort_qr)

        return master_file, fort_qr

    def _load_master_file(self) -> pd.DataFrame:
        """Загрузка мастер-файла с фильтрацией по SIZE5."""
        master_file = pd.read_excel(
            self.master_file_path,
            sheet_name=0,
            header=self.MASTER_FILE_HEADER_ROW
        )
        # Фильтруем только строки с заполненным SIZE5
        return master_file[master_file['SIZE5'].notna()]

    def _load_fort_qr_file(self) -> pd.DataFrame:
        """Загрузка файла FORT_QR с учетом специфической структуры."""
        return pd.read_excel(
            self.fort_qr_path,
            sheet_name=0,
            header=self.FORT_QR_HEADER_ROW,
            skiprows=self.FORT_QR_SKIP_ROWS
        )

    def _convert_gtin_to_string(self, master_file: pd.DataFrame, fort_qr: pd.DataFrame) -> None:
        """Преобразование GTIN кодов в строковый формат."""
        gtin_columns = ['GTIN', 'GTIN Outer', 'GTIN Case']

        for column in gtin_columns:
            if column in master_file.columns:
                master_file[column] = master_file[column].apply(
                    lambda x: str(int(x)) if pd.notna(x) else x
                )

        if 'GTIN' in fort_qr.columns:
            fort_qr['GTIN'] = fort_qr['GTIN'].apply(
                lambda x: str(int(x)) if pd.notna(x) else x
            )

    def process_data(self, master_file: pd.DataFrame, fort_qr: pd.DataFrame) -> int:
        """
        Основная обработка данных с группировкой по упаковкам.

        Args:
            master_file: Данные мастер-файла
            fort_qr: Данные файла QR-кодов

        Returns:
            Количество обработанных строк

        Raises:
            ValueError: При некорректном распределении товаров по упаковкам
        """
        self._print_matching_statistics(master_file, fort_qr)

        row_index = 0

        for _, master_row in master_file.iterrows():
            row_index = self._process_master_row(master_row, fort_qr, row_index)

        return row_index

    def _print_matching_statistics(self, master_file: pd.DataFrame, fort_qr: pd.DataFrame) -> None:
        """Вывод статистики совпадений GTIN кодов."""
        outer_matches = fort_qr[fort_qr['GTIN'].isin(master_file['GTIN Outer'])]
        case_matches = fort_qr[fort_qr['GTIN'].isin(master_file['GTIN Case'])]

        print(f"Найдено совпадений GTIN Outer: {len(outer_matches)}")
        print(f"Найдено совпадений GTIN Case: {len(case_matches)}")

    def _process_master_row(self, master_row: pd.Series, fort_qr: pd.DataFrame, row_index: int) -> int:
        """
        Обработка одной строки мастер-файла.

        Args:
            master_row: Строка из мастер-файла
            fort_qr: DataFrame с QR-кодами
            row_index: Текущий индекс строки в выходном файле

        Returns:
            Обновленный индекс строки
        """
        gtin_outer = master_row['GTIN Outer']
        gtin_case = master_row['GTIN Case']

        print(f"\nОбработка GTIN Outer: {gtin_outer}, GTIN Case: {gtin_case}")

        # Получение данных о пачках и упаковках
        pack_rows = fort_qr[fort_qr['GTIN'] == master_row['GTIN']]
        case_rows = fort_qr[fort_qr['GTIN'] == gtin_case]

        # Валидация и расчет распределения
        size5, size2 = self._get_package_sizes(master_row)
        full_chunks, remainder = self._calculate_distribution(pack_rows, size5)

        self._validate_distribution(pack_rows, case_rows, size5, size2, gtin_case, remainder)

        # Обработка полных порций
        return self._process_chunks(
            pack_rows, case_rows, master_row, full_chunks, size5, size2, row_index
        )

    def _get_package_sizes(self, master_row: pd.Series) -> tuple[int, int]:
        """Получение размеров упаковок из мастер-файла."""
        size5 = int(master_row['SIZE5']) if pd.notna(master_row['SIZE5']) else float('inf')
        size2 = int(master_row['SIZE2']) if pd.notna(master_row['SIZE2']) else float('inf')
        return size5, size2

    def _calculate_distribution(self, pack_rows: pd.DataFrame, size5: int) -> tuple[int, int]:
        """Расчет распределения пачек по коробкам."""
        total_rows = len(pack_rows)
        full_chunks = total_rows // size5
        remainder = total_rows % size5

        print(
            f"Всего пачек: {total_rows}, размер коробки (SIZE5): {size5}, "
            f"полных коробок: {full_chunks}, остаток пачек: {remainder}"
        )

        return full_chunks, remainder

    def _validate_distribution(
            self,
            pack_rows: pd.DataFrame,
            case_rows: pd.DataFrame,
            size5: int,
            size2: int,
            gtin_case: str,
            remainder: int
    ) -> None:
        """Валидация корректности распределения товаров."""
        # Проверка остатка пачек
        if remainder > 0:
            raise ValueError(
                f"Ошибка распределения для GTIN Case {gtin_case}: "
                f"{len(pack_rows)} пачек не могут быть равномерно распределены "
                f"в коробки по {size5} пачек. Остаток: {remainder} пачек."
            )

        # Проверка распределения мастер-кейсов
        total_cases = len(case_rows)
        print(f"Всего мастер-кейсов: {total_cases}, размер паллета (SIZE2): {size2}")

        if len(pack_rows) % size2 != 0:
            raise ValueError(
                f"Ошибка распределения для GTIN Case {gtin_case}: "
                f"{total_cases} мастер-кейсов не могут быть равномерно распределены "
                f"по {size2} на паллет. Остаток: {len(pack_rows) % size2} кейсов."
            )

    def _process_chunks(
            self,
            pack_rows: pd.DataFrame,
            case_rows: pd.DataFrame,
            master_row: pd.Series,
            full_chunks: int,
            size5: int,
            size2: int,
            row_index: int
    ) -> int:
        """Обработка полных порций пачек."""
        fort_qr_outer = self._get_fort_qr_data()

        for chunk_num in range(full_chunks):
            chunk = self._get_chunk(pack_rows, chunk_num, size5)

            # Получение кодов идентификации
            identification_outer = self._get_identification_code(
                fort_qr_outer, master_row['GTIN Outer'], chunk_num
            )
            identification_case = self._get_case_identification_code(
                case_rows, chunk_num, size5, size2
            )

            # Добавление строк в выходной DataFrame
            row_index = self._add_chunk_to_output(
                chunk, identification_outer, identification_case, row_index
            )

        print(f"Обработано строк для GTIN Outer {master_row['GTIN Outer']}: {row_index}")
        return row_index

    def _get_fort_qr_data(self) -> pd.DataFrame:
        """Получение данных FORT_QR для повторного использования."""
        return pd.read_excel(
            self.fort_qr_path,
            sheet_name=0,
            header=self.FORT_QR_HEADER_ROW,
            skiprows=self.FORT_QR_SKIP_ROWS
        )

    def _get_chunk(self, pack_rows: pd.DataFrame, chunk_num: int, size5: int) -> pd.DataFrame:
        """Получение порции пачек для обработки."""
        start_idx = chunk_num * size5
        end_idx = (chunk_num + 1) * size5
        return pack_rows.iloc[start_idx:end_idx]

    def _get_identification_code(
            self,
            fort_qr: pd.DataFrame,
            gtin: str,
            index: int
    ) -> str:
        """Получение кода идентификации по GTIN и индексу."""
        rows = fort_qr[fort_qr['GTIN'] == gtin]
        return rows['identificationCode'].iloc[index] if index < len(rows) else None

    def _get_case_identification_code(
            self,
            case_rows: pd.DataFrame,
            chunk_num: int,
            size5: int,
            size2: int
    ) -> str:
        """Получение кода идентификации мастер-кейса."""
        case_index = int(chunk_num // (size2 / size5))
        return case_rows['identificationCode'].iloc[case_index] if case_index < len(case_rows) else None

    def _add_chunk_to_output(
            self,
            chunk: pd.DataFrame,
            identification_outer: str,
            identification_case: str,
            row_index: int
    ) -> int:
        """Добавление порции данных в выходной DataFrame."""
        for _, pack_row in chunk.iterrows():
            self.output_df.loc[row_index] = {
                'productNameRus': pack_row['productNameRus'],
                'productNameEng': pack_row['productNameEng'],
                'identificationCode': pack_row['identificationCode'],
                'identificationCodeOuter': identification_outer,
                'identificationCodeCase': identification_case,
                'identificationCodePallet': None,
                'invoiceNo': None,
                'invoiceDate': None,
                'TotalAmount': None
            }
            row_index += 1

        return row_index

    def save_to_excel(self, row_count: int) -> None:
        """
        Сохранение обработанных данных в Excel файл.

        Args:
            row_count: Количество строк для записи
        """
        self.new_filename = self._create_new_specification_file()
        self._write_data_to_excel()

        print(f"Обработано строк: {row_count}")
        print(f"Результат сохранен в '{self.new_filename}', начиная с {self.OUTPUT_START_ROW}-й строки")

    def _create_new_specification_file(self) -> str:
        """Создание нового файла спецификации на основе шаблона."""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        new_filename = f"Invoice_Specification_{timestamp}.xlsx"
        shutil.copy(self.template_path, new_filename)
        return new_filename

    def _write_data_to_excel(self) -> None:
        """Запись данных в Excel файл."""
        wb = load_workbook(self.new_filename)
        ws = wb['Invoice specification']

        # Записываем новые данные, начиная с 11-й строки (индекс 10 в Python)
        for index, row in self.output_df.iterrows():
            excel_row = 11 + index
            ws.cell(row=excel_row, column=1, value=row['productNameRus'])
            ws.cell(row=excel_row, column=2, value=row['productNameEng'])
            ws.cell(row=excel_row, column=3, value=row['identificationCode'])
            ws.cell(row=excel_row, column=4, value=row['identificationCodeOuter'])
            ws.cell(row=excel_row, column=5, value=row['identificationCodeCase'])
            ws.cell(row=excel_row, column=6, value=row['identificationCodePallet'] if 'identificationCodePallet' in row else None)
            ws.cell(row=excel_row, column=7, value=row['invoiceNo'] if 'invoiceNo' in row else None)
            ws.cell(row=excel_row, column=8, value=row['invoiceDate'] if 'invoiceDate' in row else None)
            ws.cell(row=excel_row, column=9, value=row['TotalAmount'] if 'TotalAmount' in row else None)

        # Сохраняем изменения в новый файл
        wb.save(self.new_filename)
        # Запись данных начиная с заданной строки
        wb.close()


def generate_specification(master_file_path: str, fort_qr_path: str, template_path: str):
    """
    Функция для генерации спецификации.

    Args:
        master_file_path: Путь к мастер-файлу
        fort_qr_path: Путь к файлу QR-кодов
        template_path: Путь к шаблону
    """
    generator = SpecificationGenerator(master_file_path, fort_qr_path, template_path)

    try:
        # Загрузка и обработка данных
        master_file, fort_qr = generator.load_data()
        row_count = generator.process_data(master_file, fort_qr)
        generator.save_to_excel(row_count)

        print("Спецификация успешно создана!")
        return generator.new_filename

    except Exception as e:
        print(f"Ошибка при создании спецификации: {e}")
        return None


# Вызов класса
if __name__ == "__main__":
    # Использование функции
    result_file = generate_specification(
        master_file_path="Мастер файл номенклатуры.xlsx",
        fort_qr_path="Коды маркировки № 687 от 01.09.2025.xlsx",
        template_path="Invoice Specification template.xlsx"
    )

    if result_file:
        print(f"Файл создан: {result_file}")
